import re

import pandas as pd

from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.properties import CalcProperties  # openpyxl ≥ 3.1
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont  # openpyxl ≥ 3.1

def add_include_phrase_lookup_column(
    target_df: pd.DataFrame,
    ws_target,                     # openpyxl worksheet to modify
    no_context: pd.DataFrame,
    nc_sheet_name: str = "No Context",
    use_xlookup: bool = False,     # True -> XLOOKUP, False -> INDEX/MATCH
) -> None:
    """
    Append an 'include_phrase' formula column to ws_target that pulls from the
    No Context sheet's include_phrase by matching BOTH:
      - phrase_num  (target vs no_context)
      - phrase text (target 'phrase' header vs no_context 'phrase' header)
    The target 'phrase' column is located by header name (case-insensitive).
    """
    # Validations (dataframes)
    for col in ("phrase_num",):
        if col not in target_df.columns:
            raise ValueError(f"target_df is missing required column '{col}'.")
        if col not in no_context.columns:
            raise ValueError(f"no_context is missing required column '{col}'.")
    if "include_phrase" not in no_context.columns:
        raise ValueError("no_context is missing required column 'include_phrase'.")
    if "phrase" not in no_context.columns:
        raise ValueError("no_context is missing required column 'phrase'.")

    t_rows = len(target_df)
    t_start, t_end = 2, 1 + t_rows  # headers are on row 1

    # Create/label the new column at the end of target
    new_col_idx = target_df.shape[1] + 1
    new_col_letter = get_column_letter(new_col_idx)
    ws_target[f"{new_col_letter}1"] = "include_phrase"

    if t_rows == 0:
        return

    # Locate target 'phrase_num' and 'phrase' columns
    t_phrase_num_idx = target_df.columns.get_loc("phrase_num") + 1
    t_phrase_num_col = get_column_letter(t_phrase_num_idx)

    # Find the header named 'phrase' on the target sheet (row 1), case-insensitive
    target_phrase_text_col_letter = None
    for idx, cell in enumerate(ws_target[1], start=1):
        val = (cell.value or "").strip()
        if val and val.lower() == "phrase":
            target_phrase_text_col_letter = get_column_letter(idx)
            break
    if not target_phrase_text_col_letter:
        raise ValueError("Couldn't find a column with header 'phrase' on the target sheet.")

    # Build no_context ranges (bounded to its current size)
    nc_rows = len(no_context)
    nc_start, nc_end = 2, 1 + nc_rows
    if nc_rows == 0:
        for r in range(t_start, t_end + 1):
            ws_target[f"{new_col_letter}{r}"] = ""
        return

    # Column letters in no_context
    nc_num_col = get_column_letter(no_context.columns.get_loc("phrase_num") + 1)
    nc_txt_col = get_column_letter(no_context.columns.get_loc("phrase") + 1)
    nc_inc_col = get_column_letter(no_context.columns.get_loc("include_phrase") + 1)

    ncq = f"'{nc_sheet_name}'"  # quote sheet name for spaces
    rng_num = f"{ncq}!${nc_num_col}${nc_start}:${nc_num_col}${nc_end}"
    rng_txt = f"{ncq}!${nc_txt_col}${nc_start}:${nc_txt_col}${nc_end}"
    rng_inc = f"{ncq}!${nc_inc_col}${nc_start}:${nc_inc_col}${nc_end}"

    # Per-row formulas with 2 criteria
    for r in range(t_start, t_end + 1):
        a_num = f"{t_phrase_num_col}{r}"                 # current row phrase_num
        a_txt = f"${target_phrase_text_col_letter}{r}"   # current row phrase text (by header)

        if use_xlookup:
            # =IFERROR(XLOOKUP(1, (rng_num=A2)*(rng_txt=$<phrase_col>2), rng_inc), "")
            formula = (
                f'=IFERROR('
                f'XLOOKUP(1, ({rng_num}={a_num})*({rng_txt}={a_txt}), {rng_inc})'
                f', "")'
            )
        else:
            # =IFERROR(INDEX(rng_inc, MATCH(1, INDEX((rng_num=A2)*(rng_txt=$<phrase_col>2),0),0)), "")
            formula = (
                f'=IFERROR('
                f'INDEX({rng_inc}, MATCH(1, INDEX(({rng_num}={a_num})*({rng_txt}={a_txt}), 0), 0))'
                f', "")'
            )

        ws_target[f"{new_col_letter}{r}"] = formula

        
def add_llr_metrics(
    ws_llr, *,
    llr_sheet_name: str,
    nc_sheet_name: str,
    known_sheet_name: str,
    unknown_sheet_name: str,
) -> None:
    LLR = f"'{llr_sheet_name}'"
    NC  = f"'{nc_sheet_name}'"
    KN  = f"'{known_sheet_name}'"
    UN  = f"'{unknown_sheet_name}'"

    # headers
    ws_llr["D1"] = "num_phrases"
    ws_llr["E1"] = "phrases_kept"
    ws_llr["F1"] = "pmf_no_context"
    ws_llr["G1"] = "pmf_known"
    ws_llr["H1"] = "pmf_unknown"
    ws_llr["I1"] = "llr_no_context"
    ws_llr["J1"] = "llr_known"
    ws_llr["K1"] = "llr_unknown"

    last_row = ws_llr.max_row
    if last_row < 2:
        return

    for r in range(2, last_row + 1):
        a = f"A{r}"  # phrase_num
        b = f"B{r}"  # phrase_occurence

        # D: num_phrases
        ws_llr[f"D{r}"] = f"=COUNTIFS({NC}!$A:$A,{LLR}!${a})"

        # E: phrases_kept (include_phrase in NC col J)
        ws_llr[f"E{r}"] = f"=COUNTIFS({NC}!$A:$A,{LLR}!${a},{NC}!$J:$J,TRUE)"

        # F: pmf_no_context  (add NC J:J TRUE to both parts)
        ws_llr[f"F{r}"] = (
            f"=IFERROR("
            f"SUMIFS({NC}!$H:$H,"
            f"{NC}!$A:$A,{LLR}!${a},"
            f"{NC}!$C:$C,\"reference\","
            f"{NC}!$J:$J,TRUE)"
            f"/"
            f"SUMIFS({NC}!$H:$H,"
            f"{NC}!$A:$A,{LLR}!${a},"
            f"{NC}!$J:$J,TRUE)"
            f",0)"
        )

        # G: pmf_known  (add KN N:N TRUE to both parts)  — also filters C='reference' in D:D per your spec
        ws_llr[f"G{r}"] = (
            f"=IFERROR("
            f"SUMIFS({KN}!$K:$K,"
            f"{KN}!$A:$A,{LLR}!${a},"
            f"{KN}!$B:$B,{LLR}!${b},"
            f"{KN}!$D:$D,\"reference\","
            f"{KN}!$N:$N,TRUE)"
            f"/"
            f"SUMIFS({KN}!$K:$K,"
            f"{KN}!$A:$A,{LLR}!${a},"
            f"{KN}!$B:$B,{LLR}!${b},"
            f"{KN}!$N:$N,TRUE)"
            f",0)"
        )

        # H: pmf_unknown  (add UN N:N TRUE to both parts)
        ws_llr[f"H{r}"] = (
            f"=IFERROR("
            f"SUMIFS({UN}!$K:$K,"
            f"{UN}!$A:$A,{LLR}!${a},"
            f"{UN}!$B:$B,{LLR}!${b},"
            f"{UN}!$D:$D,\"reference\","
            f"{UN}!$N:$N,TRUE)"
            f"/"
            f"SUMIFS({UN}!$K:$K,"
            f"{UN}!$A:$A,{LLR}!${a},"
            f"{UN}!$B:$B,{LLR}!${b},"
            f"{UN}!$N:$N,TRUE)"
            f",0)"
        )

        # I/J/K: base-10 logs of 1/pmf*
        ws_llr[f"I{r}"] = f"=IFERROR(LOG(1/F{r},10),0)"
        ws_llr[f"J{r}"] = f"=IFERROR(LOG(1/G{r},10),0)"
        ws_llr[f"K{r}"] = f"=IFERROR(LOG(1/H{r},10),0)"

def add_metadata_metrics(ws_meta, *, llr_sheet_name: str, ws_llr) -> None:
    """
    metadata!J: num_phrases        = distinct count of nonblank LLR!C
    metadata!K: phrases_kept       = distinct count of LLR!C where LLR!E > 0
    metadata!L: llr_no_context     = SUMIFS(LLR!I:I, LLR!B:B, 1)
    metadata!M: llr_known          = SUM(LLR!J:J)
    metadata!N: llr_unknown        = SUM(LLR!K:K)
    metadata!O: normalised_llr_no_context = L2 / K2 (IFERROR -> 0)
    metadata!P: normalised_llr_known      = M2 / K2 (IFERROR -> 0)
    metadata!Q: normalised_llr_unknown    = N2 / K2 (IFERROR -> 0)
    """
    LLR = f"'{llr_sheet_name}'"
    last_r = max(ws_llr.max_row, 2)

    c_rng = f"{LLR}!$C$2:$C${last_r}"
    e_rng = f"{LLR}!$E$2:$E${last_r}"

    # J: all distinct phrases (nonblank) — legacy SUMPRODUCT/COUNTIF pattern
    ws_meta["J1"] = "num_phrases"
    ws_meta["J2"] = f"=SUMPRODUCT(({c_rng}<>\"\")/COUNTIF({c_rng},{c_rng}))"

    # K: distinct phrases where any row for that phrase has E>0
    ws_meta["K1"] = "phrases_kept"
    ws_meta["K2"] = (
        f"=SUMPRODUCT("
        f"({c_rng}<>\"\")*"
        f"(COUNTIFS({c_rng},{c_rng},{e_rng},\">0\")>0)"
        f"/COUNTIF({c_rng},{c_rng})"
        f")"
    )

    # L/M/N: sums of LLR columns with stated conditions
    ws_meta["L1"] = "llr_no_context"
    ws_meta["L2"] = f"=SUMIFS({LLR}!$I:$I,{LLR}!$B:$B,1)"   # only rows with occurrence = 1

    ws_meta["M1"] = "llr_known"
    ws_meta["M2"] = f"=SUM({LLR}!$J:$J)"

    ws_meta["N1"] = "llr_unknown"
    ws_meta["N2"] = f"=SUM({LLR}!$K:$K)"

    # O/P/Q: normalized by phrases_kept (metadata!K2)
    ws_meta["O1"] = "normalised_llr_no_context"
    ws_meta["O2"] = "=IFERROR(L2/K2,0)"

    ws_meta["P1"] = "normalised_llr_known"
    ws_meta["P2"] = "=IFERROR(M2/K2,0)"

    ws_meta["Q1"] = "normalised_llr_unknown"
    ws_meta["Q2"] = "=IFERROR(N2/K2,0)"

def style_entire_workbook(wb, *, header_font="Calibri", header_size=11,
                          header_grey="F2F2F2", freeze_headers=True):
    """
    Apply uniform header style + thin borders to all cells across all sheets.
    - Header row assumed at row 1.
    - Borders applied to the entire used range (1..max_row, 1..max_column).
    """
    # Build reusable styles (docs: Font/Fill/Border/Alignment/NamedStyle)
    # https://openpyxl.readthedocs.io/en/3.1/styles.html
    thin = Side(style="thin", color="000000")
    all_borders = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_style = NamedStyle(name="__hdr__auto")
    header_style.font = Font(name=header_font, size=header_size, bold=True)
    header_style.fill = PatternFill("solid", fgColor=header_grey)
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = all_borders

    # Register style once (ignore if it already exists)
    if "__hdr__auto" not in wb.named_styles:
        wb.add_named_style(header_style)

    for ws in wb.worksheets:
        max_r, max_c = ws.max_row or 1, ws.max_column or 1

        # 1) Header row styling (row 1)
        for c in range(1, max_c + 1):
            cell = ws.cell(row=1, column=c)
            # If a cell existed, apply the named style; if it's empty, this is still fine
            cell.style = "__hdr__auto"

        # 2) Borders for all cells in used range (headers + data)
        # Pattern: iterate rows and apply Border to each cell
        # https://stackoverflow.com/a/18844061  (range borders with openpyxl)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                ws.cell(row=r, column=c).border = all_borders

        # 3) Optional: freeze header row
        # set the pane to first cell BELOW headers, i.e. A2
        # https://automatetheboringstuff.com/2e/chapter13/
        if freeze_headers:
            ws.freeze_panes = "A2"

def reorder_sheets(wb, desired_order: list[str]) -> None:
    """
    Reorder workbook sheets to match desired_order (by title).
    Any sheets not listed are appended at the end, preserving their relative order.
    """
    # Keep only existing sheet objects in the requested order
    ordered = [wb[name] for name in desired_order if name in wb.sheetnames]
    # Append any other sheets not specified
    ordered += [ws for ws in wb.worksheets if ws.title not in desired_order]
    # Assign back (openpyxl supports reassigning the private _sheets list)
    wb._sheets = ordered

def wrap_text_in_docs(ws_docs, cols=("A", "B")) -> None:
    """
    Turn on text wrapping for the given columns (default: A & B) on the docs sheet.
    """
    wrap = Alignment(wrapText=True, horizontal="left", vertical="top")
    max_r = ws_docs.max_row or 1
    for col in cols:
        for r in range(1, max_r + 1):
            cell = ws_docs[f"{col}{r}"]
            # Rebind alignment (styles are immutable)
            cell.alignment = wrap

def _header_map(ws):
    m = {}
    for i, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            k = str(cell.value).strip().lower()
            if k and k not in m:
                m[k] = i
    return m

def _normalize_cols(ws, cols):
    if cols is None:
        return list(range(1, ws.max_column + 1))
    lookup = _header_map(ws)
    out = []
    for c in cols:
        if isinstance(c, int):
            out.append(c)
        else:
            s = str(c).strip()
            if s.lower() in lookup:
                out.append(lookup[s.lower()])
            else:
                out.append(column_index_from_string(s))  # may raise if bad
    return out

def _is_number(v):
    return isinstance(v, (int, float, Decimal)) and not isinstance(v, bool)

def _is_date(v):
    return isinstance(v, (date, datetime))

def autofit_columns(
    ws,
    cols=None,                  # None = all; or header names / letters / indexes
    min_width=8,
    max_width=60,               # hard cap for general/text columns
    numeric_max_width=14,       # tighter cap for numeric columns
    date_max_width=12,          # tighter cap for dates
    bool_max_width=6,           # tiny for TRUE/FALSE
    padding=2.0,
    wrap=True,
    shrink=False,
    ignore_outlier_pct=0.98,    # ignore top 2% lengths
    sample_rows=500,            # sample up to N rows to detect type & lengths
):
    col_indexes = _normalize_cols(ws, cols)
    for col_idx in col_indexes:
        letter = get_column_letter(col_idx)

        # --- sample cells to detect type and measure lengths
        lengths = []
        n_numeric = n_date = n_bool = n_total = 0

        # iterate with values_only for speed; sample up to sample_rows
        max_r = ws.max_row
        end_r = min(max_r, sample_rows)
        for row in ws.iter_rows(min_row=2, max_row=end_r,  # skip header for typing
                                min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            n_total += 1
            if _is_number(v):
                n_numeric += 1
                # estimate displayed length: digits + commas + decimal part
                s = f"{v:,}" if float(v).is_integer() else f"{v:,.2f}"
            elif _is_date(v):
                n_date += 1
                s = v.strftime("%Y-%m-%d %H:%M") if isinstance(v, datetime) else v.strftime("%Y-%m-%d")
            elif isinstance(v, bool):
                n_bool += 1
                s = "TRUE" if v else "FALSE"
            else:
                s = str(v)

            # longest visual line if there are line breaks
            parts = s.splitlines() if "\n" in s else [s]
            for part in parts:
                # crude width heuristic; non-ASCII a bit wider
                L = sum(1.2 if ord(ch) > 255 else 1.0 for ch in part)
                lengths.append(L)

        # include header
        header_val = ws.cell(row=1, column=col_idx).value
        if header_val:
            lengths.append(len(str(header_val)) * 1.05)

        # decide the dominant type in this column
        dominant = None
        if n_total:
            ratios = [(n_numeric, "num"), (n_date, "date"), (n_bool, "bool")]
            dominant = max(ratios, key=lambda t: t[0])[1] if max(ratios)[0] / n_total >= 0.6 else None

        # pick a base length with percentile cap
        if lengths:
            lengths.sort()
            k = int(len(lengths) * ignore_outlier_pct) - 1
            k = max(0, min(k, len(lengths) - 1))
            base = lengths[k]
        else:
            base = min_width

        # choose cap by type
        cap = {
            "num": numeric_max_width,
            "date": date_max_width,
            "bool": bool_max_width,
            None: max_width,
        }[dominant]

        width = min(cap, max(min_width, base + padding))
        ws.column_dimensions[letter].width = width

        # apply wrap / shrink (rebind style)
        if wrap or shrink:
            aln = Alignment(
                wrapText=wrap if wrap else None,
                shrinkToFit=shrink if shrink else None,
                vertical="top"
            )
            for r in range(1, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).alignment = aln

def highlight_overlaps_in_docs(
    wb,
    docs_sheet="docs",
    llr_sheet="LLR",
    phrase_col_letter="C",
    text_cells=("A5", "B5"),
    color_single="C00000",      # red for single matches
    color_overlap="FF8C00",     # orange for overlaps
    make_bold=True,
    case_insensitive=True,
    whole_words=False,          # set True to require word boundaries
):
    ws_docs = wb[docs_sheet]
    ws_llr  = wb[llr_sheet]

    # 1) distinct nonblank phrases from LLR!C (row 2+)
    col_idx = ord(phrase_col_letter.upper()) - 64
    phrases = []
    for (v,) in ws_llr.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
        if v is None:
            continue
        s = str(v).strip()
        if s:
            phrases.append(s)
    phrases = sorted(set(phrases), key=lambda s: (-len(s), s))  # longest-first
    if not phrases:
        return

    # 2) big alternation; use lookahead to get overlapping matches
    #    (?=(...)) finds zero-width positions; group(1) is the actual phrase
    inner = "|".join(re.escape(p) for p in phrases)
    if whole_words:
        inner = r"\b(?:" + inner + r")\b"
    pattern = r"(?=(" + inner + r"))"
    flags = re.IGNORECASE if case_insensitive else 0
    rx = re.compile(pattern, flags)

    def make_runs(text: str) -> CellRichText:
        if not text:
            return CellRichText([""])

        # collect all (start, end) spans using the lookahead
        spans = []
        for m in rx.finditer(text):
            start = m.start()
            end = start + len(m.group(1))
            spans.append((start, end))
        if not spans:
            return CellRichText([text])

        # coverage map: how many phrases cover each character
        cov = [0] * len(text)
        for a, b in spans:
            for i in range(a, b):
                cov[i] += 1

        # split into runs by coverage transitions
        runs, i = [], 0
        while i < len(text):
            start = i
            level = cov[i] if i < len(cov) else 0
            i += 1
            while i < len(text) and (cov[i] if i < len(cov) else 0) == level:
                i += 1
            seg = text[start:i]
            if level == 0:
                runs.append(seg)  # plain
            elif level == 1:
                runs.append(TextBlock(InlineFont(color=color_single, b=make_bold), seg))
            else:
                runs.append(TextBlock(InlineFont(color=color_overlap, b=make_bold), seg))
        return CellRichText(runs)

    # 3) apply to the requested docs cells
    for addr in text_cells:
        raw = "" if ws_docs[addr].value is None else str(ws_docs[addr].value)
        ws_docs[addr].value = make_runs(raw)

def create_excel_template(
    known: pd.DataFrame,
    unknown: pd.DataFrame,
    no_context: pd.DataFrame,
    metadata: pd.DataFrame,
    docs: pd.DataFrame,
    path: str | Path = "template.xlsx",
    known_sheet: str = "known",
    unknown_sheet: str = "unknown",
    nc_sheet: str = "no context",
    metadata_sheet: str = "metadata",
    docs_sheet: str = "docs",
    llr_sheet: str = "LLR",
    use_xlookup: bool = False,
    highlight_phrases: bool = False
) -> Path:
    """
    Writes all sheets, builds a distinct phrases 'LLR' table, adds include_phrase lookups
    to Known & Unknown, and then adds your LLR formulas (D..H).
    """
    path = Path(path)

    # Preconditions
    for name, df in [("known", known), ("unknown", unknown), ("no_context", no_context)]:
        if "phrase_num" not in df.columns:
            raise ValueError(f"{name} is missing required column 'phrase_num'.")

    # Ensure include flags on no_context
    no_context = no_context.copy()
    no_context["include_phrase"] = True

    # Create LLR table (distinct phrases)
    llr_cols = ['phrase_num', 'phrase_occurence', 'original_phrase']
    distinct_phrases = (
        pd.concat([unknown[llr_cols], known[llr_cols]], ignore_index=True)
        .drop_duplicates()
        .sort_values(['phrase_num', 'phrase_occurence'], kind='mergesort')
        .reset_index(drop=True)
    )

    # Choose writer mode safely
    writer_mode = "a" if path.exists() else "w"
    writer_kwargs = {"engine": "openpyxl", "mode": writer_mode}
    if writer_mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"  # only valid in append mode
        

    with pd.ExcelWriter(path, **writer_kwargs) as writer:
        # Write sheets
        docs.to_excel(writer, index=False, sheet_name=docs_sheet)
        known.to_excel(writer, index=False, sheet_name=known_sheet)
        unknown.to_excel(writer, index=False, sheet_name=unknown_sheet)
        no_context.to_excel(writer, index=False, sheet_name=nc_sheet)
        distinct_phrases.to_excel(writer, index=False, sheet_name=llr_sheet)
        metadata.to_excel(writer, index=False, sheet_name=metadata_sheet)

        # Add formulas to Known/Unknown
        wb = writer.book
        ws_meta = wb[metadata_sheet]
        _ = wb[nc_sheet]  # assert exists
        ws_known   = wb[known_sheet]
        ws_unknown = wb[unknown_sheet]
        ws_llr     = wb[llr_sheet]
        ws_docs = wb[docs_sheet]

        # Sometimes highlighting fails so adding for now
        if highlight_phrases:
            highlight_overlaps_in_docs(
                wb,
                docs_sheet=docs_sheet,   # whatever you named it
                llr_sheet=llr_sheet,     # whatever you named it
                color_single="C00000",
                color_overlap="FF8C00",
                make_bold=True,
                case_insensitive=True,
            )
        
        add_include_phrase_lookup_column(known, ws_known, no_context, nc_sheet, use_xlookup)
        add_include_phrase_lookup_column(unknown, ws_unknown, no_context, nc_sheet, use_xlookup)

        # Add LLR metrics (your D..H columns)
        add_llr_metrics(
            ws_llr,
            llr_sheet_name=llr_sheet,
            nc_sheet_name=nc_sheet,
            known_sheet_name=known_sheet,
            unknown_sheet_name=unknown_sheet,
        )

        # Add the metadata metrics
        add_metadata_metrics(ws_meta, llr_sheet_name=llr_sheet, ws_llr=ws_llr)
        
        # Style every sheet
        style_entire_workbook(wb)

        wrap_text_in_docs(ws_docs, cols=("A","B"))
        autofit_columns(ws_docs)
        
        # Reorder sheets exactly as requested
        reorder_sheets(
            wb,
            desired_order=[docs_sheet, metadata_sheet, nc_sheet, known_sheet, unknown_sheet, llr_sheet]
        )

        # Force Excel to do a full calc on open (newer openpyxl API)
        wb.calculation = CalcProperties(fullCalcOnLoad=True)

    return path
